# _*_coding:utf-8
# @Time      :2023/11/23 0:18
# @Author    :DQ
# @FileName: ua.py

import openpyxl
import requests as res
from fake_useragent import UserAgent
import re
import os


class Ua:
    # 类的默认URL
    url = 'https://m.acxsw.com/'

    def set_url(self, new_url):
        # 设置类的新URL的方法
        self.url = new_url

    def fake_ua_get(self):
        # 使用随机User-Agent进行GET请求的方法
        ua = UserAgent()
        head = {
            'User-Agent': ua.random
        }
        rs = res.get(self.url, headers=head, stream=True)
        return rs

    def get_h1_title(self):
        # 从HTML中提取并返回第一个h1标签内的内容的方法
        html = self.fake_ua_get().text
        pattern = r'<h1>(.*?)</h1>'
        match = re.search(pattern, html)
        if match:
            title = match.group(1)
        return title

    def download_html(self, title):
        # 下载HTML内容到文件的方法
        html = self.fake_ua_get().text
        path = f'C:/Users/86132/PycharmProjects/pythonProject/Acxiaoshuo/download/html/{title}.html'

        if os.path.exists(path):
            # 如果文件存在，检查内容是否相同，如果不同，创建一个带有递增索引的新文件
            with open(path, mode='r', encoding='utf-8') as f:
                existing_content = f.read()
            pattern = r'<a href="(.*?)">(.*?)</a>'
            match_1 = re.search(pattern, html)
            match_2 = re.search(pattern, existing_content)
            author_1 = match_1.group(2)
            author_2 = match_2.group(2)
            if author_1 == author_2:
                print('文件已存在且内容相同！')
                return False
            else:
                index = 1
                path = f'C:/Users/86132/PycharmProjects/pythonProject/Acxiaoshuo/download/html/{title}_{index}.html'
                index += 1

        try:
            with open(path, mode='w', encoding='utf-8') as f:
                f.write(html)
                return True
        except  Exception as e:
            print(f"错误：{e}")
            return False

    def read_html(self, dir, title):
        try:
            with open(f'C:/Users/86132/PycharmProjects/pythonProject/Acxiaoshuo/download/html/{dir}{title}.html',
                      mode='r', encoding='utf-8') as f:
                html_content = f.read()
                return html_content

        except  Exception as e:
            print(f"错误：{e}")

    def read_column_data(self, file_path, column_index):
        # 打开工作簿
        wb = openpyxl.load_workbook(file_path)

        # 获取活动的工作表
        sheet = wb.active

        # 获取指定列的数据
        column_data = []
        for i in range(2, sheet.max_row + 1):
            column_data.append(sheet.cell(row=i, column=column_index).value)

        # 关闭工作簿
        wb.close()

        return column_data

    # 从第二行开始，以列为标准添加数据
    def date_split(self, values, file_path, column):
        # 打开excel文件
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        for index, value in enumerate(values, start=2):
            sheet.cell(row=index, column=column, value=value)
        # 关闭工作簿
        wb.close()
        # 保存工作簿
        wb.save(file_path)
        print('保存完成。')

    def download_img(self, save_path):
        # 检查文件是否已经存在
        if os.path.exists(save_path):
            # 如果存在，生成一个新的文件名
            base, extension = os.path.splitext(save_path)
            index = 1
            while os.path.exists(f"{base}_{index}{extension}"):
                index += 1
            save_path = f"{base}_{index}{extension}"

        # 使用请求头访问
        response = self.fake_ua_get()
        if response.status_code == 200:
            with open(save_path, 'wb') as file:
                for chunk in response.iter_content(1024):
                    file.write(chunk)
            print(f"图片已成功下载到 {save_path}")
        else:
            print(f"下载失败，HTTP 状态码: {response.status_code}")
