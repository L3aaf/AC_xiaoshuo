# _*_coding:utf-8
# @Time      :2023/11/25 12:39
# @Author    :DQ
# @FileName: god_img.py
import os
import time

import openpyxl

import Acxiaoshuo.tool.ua
from lxml import html


class God_img:

    def get_srcs_with_alts(self):
        # 实例化一个名为tf的工具对象，用于分析网页内容
        tf = Acxiaoshuo.tool.isTureF.Tf()
        # 设置工具对象的目录为'god/'
        tf.set_dir('god/')
        # 获取标题列表
        title_list = tf.get_title()

        # 实例化一个名为ua的用户代理对象
        ua = Acxiaoshuo.tool.ua.Ua()
        # 初始化空的图片链接列表和图片描述列表
        src_list = []
        alt_list = []

        # 遍历标题列表
        for title in title_list:
            # 读取god/目录下指定标题的HTML内容
            html_content = ua.read_html('god/', title)
            # 使用XPath解析HTML内容为树形结构
            tree = html.fromstring(html_content)
            # 获取图片链接列表
            src_values = tree.xpath('/html/body/div[2]/section/div[1]/img/@src')
            # 获取图片描述列表
            alt_values = tree.xpath('/html/body/div[2]/section/div[1]/img/@alt')
            # 将图片链接列表和图片描述列表添加到对应的列表中
            src_list.append(src_values)
            alt_list.append(alt_values)

        # 返回图片链接列表和图片描述列表
        return src_list, alt_list


    def input_date(self):
        # 获取所有图片链接列表和图片描述列表
        all_list = self.get_srcs_with_alts()
        src_list, alt_list = all_list
        # 设置文件路径为指定位置
        file_path = 'C:/Users/86132/PycharmProjects/pythonProject/Acxiaoshuo/download/html/excel/大神小说数据.xlsx'
        # 打开已存在的工作簿
        wb = openpyxl.load_workbook(file_path)
        # 获取活动的工作表
        sheet = wb.active
        # 在第一行的第10列添加标题'图片地址链接'
        sheet.cell(row=1, column=10, value='图片地址链接')

        # 遍历图片链接列表，将链接字符串添加到工作表的第10列
        for index, link in enumerate(src_list, start=2):
            src_values_str = ' '.join(link)
            sheet.cell(row=index, column=10, value=src_values_str)
        # 关闭工作簿
        wb.close()
        # 保存工作簿
        wb.save(file_path)

    def download_all_img(self):
        # 获取所有图片链接列表和图片描述列表
        all_list = self.get_srcs_with_alts()
        src_list, alt_list = all_list
        # 实例化一个名为ua的用户代理对象
        ua = Acxiaoshuo.tool.ua.Ua()

        # 遍历图片链接列表和图片描述列表
        for src, alt in zip(src_list, alt_list):
            # 设置用户代理对象的URL为当前图片链接
            ua.set_url(src[0])
            # 设置保存路径为指定位置，文件名为图片描述+'.jpg'
            save_path = os.path.join(r'C:\Users\86132\PycharmProjects\pythonProject\Acxiaoshuo\download\html\img\god',
                                     f'{alt[0]}.jpg')
            # 下载图片到指定路径
            ua.download_img(save_path)
            # 反爬，等待1秒
            time.sleep(1)
            print('下载中---')
        # 打印下载完成的提示
        print('下载完成！')