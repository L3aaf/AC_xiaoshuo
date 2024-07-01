# _*_coding:utf-8
# @Time      :2023/11/23 0:28
# @Author    :DQ
# @FileName: hot.py


import re

import Acxiaoshuo.tool.ua
import Acxiaoshuo.tool.isTureF
from lxml import html
import openpyxl


class hot:

    def get_links(self):
        # 创建 Ua 对象
        ua = Acxiaoshuo.tool.ua.Ua()

        # 通过 Ua 对象获取 HTML 内容
        html_content = ua.read_html('', 'AC小说_最值得书友收藏的网络小说阅读网')

        # 使用 lxml 库解析 HTML 内容
        tree = html.fromstring(html_content)

        # 使用 XPath 获取排序链接和标题
        sort_links = tree.xpath(r'/html/body/div[2]/aside/ul/li[/]/a[1]/@href')
        sort_title = tree.xpath(r'/html/body/div[2]/aside/ul/li[/]/a[1]/text()')

        # 存储链接和标题的列表
        links_list = []
        title_list = []

        # 遍历标题和链接，并处理
        for title, link in zip(sort_title, sort_links):
            # 拼接完整链接
            link = f'https://m.acxsw.com{link}'

            # 将链接添加到列表
            links_list.append(link)
            title_list.append(title)

        return links_list, title_list

    def get_content(self):
        titles_list = self.get_links()[1]
        ua = Acxiaoshuo.tool.ua.Ua()

        # 创建工作簿（Workbook）
        wb = openpyxl.Workbook()

        # 获取默认工作表（Active Sheet）
        sheet = wb.active

        # 添加标题行
        sheet.append(
            ['书名', '作者', '频道', '总字数', '状态', '最新章节', '概要', '最后更新时间(截止至2023.11.23)',
             '原文链接',
             '图片地址链接']
        )

        file_path = 'C:/Users/86132/PycharmProjects/pythonProject/Acxiaoshuo/download/html/excel/热点数据.xlsx'

        for index, title_list in enumerate(titles_list, start=2):
            html_content = ua.read_html('hot/', title_list)
            tree = html.fromstring(html_content)

            titles = tree.xpath(r'/html/body/div[2]/section/div[1]/div/h1/text()')
            contents = tree.xpath(r'//*[@id="info"]/div[1]/p[2]/text()')
            contents = contents if contents else ['未提供']

            authors = tree.xpath(r'/html/body/div[2]/section/div[1]/div/i/a/text()')
            sortvisits = tree.xpath(r'/html/body/div[2]/section/div[1]/div/p[1]/span[1]/text()')
            counts = tree.xpath(r'/html/body/div[2]/section/div[1]/div/p[1]/span[2]/text()')
            states = tree.xpath(r'/html/body/div[2]/section/div[1]/div/p[1]/span[3]/text()')
            newests = tree.xpath(r'/html/body/div[2]/section/div[1]/div/div[1]/a/text()')
            newest_times = tree.xpath(r'/html/body/div[2]/section/div[1]/div/div[1]/em/text()')

            for title, content, author, sortvisit, count, state, newest, newest_time in \
                    zip(titles, contents, authors, sortvisits, counts, states, newests, newest_times):
                sheet.cell(row=index, column=1, value=title)
                sheet.cell(row=index, column=2, value=author)
                sheet.cell(row=index, column=3, value=sortvisit)
                sheet.cell(row=index, column=4, value=count)
                sheet.cell(row=index, column=5, value=state)
                sheet.cell(row=index, column=6, value=newest)
                sheet.cell(row=index, column=7, value=content)
                sheet.cell(row=index, column=8, value=newest_time)
            print("循环完成。 保存到 Excel...")
            print("文件路径:", file_path)
            print("Index:", index)
        wb.save(file_path)

    def get_url_with_title(self):
        titles_list = self.get_links()[1]
        ua = Acxiaoshuo.tool.ua.Ua()
        file_path = 'C:/Users/86132/PycharmProjects/pythonProject/Acxiaoshuo/download/html/excel/热点数据.xlsx'
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        for index, title_list in enumerate(titles_list, start=2):
            html_content = ua.read_html('hot/', title_list)
            tree = html.fromstring(html_content)
            titles = tree.xpath(r'/html/body/div[2]/section/div[1]/div/h1/text()')
            url_pattern = re.compile(r'url=(https://[^\s"]+)')
            match = url_pattern.search(html_content)
            extracted_url = match.group(1)

            for title, url in zip(titles, extracted_url):
                text = f'{title} - {extracted_url}'
                sheet.cell(row=index, column=9, value=text)
            print("循环完成。 保存到 Excel...")
            print("文件路径:", file_path)
            print("Index:", index)

        wb.save(file_path)
